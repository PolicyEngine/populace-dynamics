"""Fetch NCHS national marriage & divorce rate trends as a reference.

The external marriage/divorce anchor for the gate-2 family-transition
floors. NCHS/NVSS publishes provisional national marriage and divorce
counts and *crude rates per 1,000 total population* (2000-latest) as a
single machine-readable xlsx linked from
``cdc.gov/nchs/nvss/marriage-divorce.htm``. The workbook stacks two tables
on one sheet -- marriages then divorces & annulments -- each a
``Year | count | population | rate`` block with descending years and
footnote rows appended.

Two documented conventions carried through verbatim:

* year cells carry trailing footnote digits (e.g. ``"20231"`` = 2023 with
  footnote 1); the leading four digits are the year;
* the divorce rate's population denominator is smaller than the marriage
  rate's because five states (California, Hawaii, Indiana, Minnesota, New
  Mexico) do not report divorces to NVSS -- this is baked into NCHS's own
  "national" divorce rate, not a fetch bug. It is recorded, not adjusted.

Because these are *crude* rates per 1,000 total population, they are a
loose order-of-magnitude anchor for the PSID age/duration-specific
hazards, not a level target -- the gate-2 floor reports the ratio with the
concept delta stated, never tuning to it.

The parsed latest-year rates are validated against the CDC FastStats
headline (2023: marriage 6.1, divorce 2.4). Written with full provenance
-- source URL, fetch timestamp, sha256 of the xlsx, the NVSS citation --
to ``data/external/nchs_marriage_divorce_rates_<year>.json``.

Run from the repository root (needs openpyxl)::

    .venv/bin/python scripts/fetch_nchs_marriage_divorce.py
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import json
import ssl
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "external"

#: Latest year in the published trend file.
LATEST_YEAR = 2023
FIRST_YEAR = 2000

_DVS = "https://www.cdc.gov/nchs/data/dvs/marriage-divorce"
XLSX_URL = f"{_DVS}/national-marriage-divorce-rates-00-23.xlsx"
PDF_URL = f"{_DVS}/national-marriage-divorce-rates-00-23.pdf"
NVSS_PAGE = "https://www.cdc.gov/nchs/nvss/marriage-divorce.htm"
FASTATS_PAGE = "https://www.cdc.gov/nchs/fastats/marriage-divorce.htm"

NVSS_CITATION = (
    "CDC/NCHS, National Vital Statistics System. National Marriage and "
    "Divorce Rate Trends for 2000-2023 (provisional). Hyattsville, MD: "
    "National Center for Health Statistics. Available from: "
    "https://www.cdc.gov/nchs/nvss/marriage-divorce.htm."
)

#: CDC FastStats headline crude rates per 1,000 total population, latest
#: three years; the parsed feed must match these exactly.
PUBLISHED_RATES: dict[int, dict[str, float]] = {
    2023: {"marriage": 6.1, "divorce": 2.4},
    2022: {"marriage": 6.2, "divorce": 2.4},
    2021: {"marriage": 6.0, "divorce": 2.5},
}
#: States excluded from the national divorce rate's denominator.
DIVORCE_EXCLUDED_STATES = (
    "California",
    "Hawaii",
    "Indiana",
    "Minnesota",
    "New Mexico",
)


def _ssl_context() -> ssl.SSLContext:
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl.create_default_context()


def _fetch_bytes(url: str) -> bytes:
    request = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (populace-dynamics fetch)"}
    )
    with urllib.request.urlopen(
        request, timeout=120, context=_ssl_context()
    ) as response:
        return response.read()


def _year_of(cell: object) -> int | None:
    """Leading four-digit year of a data-row label (footnotes stripped)."""
    text = str(cell).strip()
    if len(text) >= 4 and text[:4].isdigit():
        year = int(text[:4])
        if FIRST_YEAR <= year <= LATEST_YEAR:
            return year
    return None


def _parse_block(
    rows: list[tuple], header_idx: int
) -> dict[int, dict[str, float]]:
    """Read a ``Year | count | population | rate`` block below its header."""
    out: dict[int, dict[str, float]] = {}
    for raw in rows[header_idx + 1 :]:
        if not raw or raw[0] is None:
            continue
        year = _year_of(raw[0])
        if year is None:
            break  # first non-year row ends the block (footnotes)
        out[year] = {
            "count": int(raw[1]),
            "population": int(raw[2]),
            "rate_per_1000": float(raw[3]),
        }
    return out


def parse_workbook(
    xlsx_bytes: bytes,
) -> dict[str, dict[int, dict[str, float]]]:
    """Parse the two stacked tables into ``{"marriage"|"divorce": {year:...}}``.

    The two ``Year`` header rows are located by their leading cell and the
    second column's label (``Marriages`` vs ``Divorces``), so a reordered
    workbook fails loudly rather than silently swapping the series.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    sheet = wb[wb.sheetnames[0]]
    rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    wb.close()

    headers = [
        i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Year"
    ]
    if len(headers) < 2:
        raise ValueError(
            f"expected two 'Year' header rows, found {len(headers)}; the "
            "NCHS workbook layout may have changed."
        )
    labels = {i: str(rows[i][1]).lower() for i in headers}
    marriage_idx = next(i for i in headers if "marriage" in labels[i])
    divorce_idx = next(i for i in headers if "divorce" in labels[i])
    return {
        "marriage": _parse_block(rows, marriage_idx),
        "divorce": _parse_block(rows, divorce_idx),
    }


def validate(tables: dict[str, dict[int, dict[str, float]]]) -> dict[str, Any]:
    """Check latest-three-year rates against the FastStats headline."""
    for year, published in PUBLISHED_RATES.items():
        for series, rate in published.items():
            got = tables[series].get(year)
            if got is None:
                raise ValueError(f"{series} {year}: missing from the feed.")
            if round(got["rate_per_1000"], 1) != rate:
                raise ValueError(
                    f"{series} {year}: parsed rate {got['rate_per_1000']} "
                    f"!= published {rate}."
                )
    # Both series present for the full advertised span, and marriage rate
    # exceeds divorce rate every year (a stable qualitative fact).
    years = sorted(set(tables["marriage"]) & set(tables["divorce"]))
    if years[0] != FIRST_YEAR or years[-1] != LATEST_YEAR:
        raise ValueError(
            f"year span {years[0]}-{years[-1]} != "
            f"{FIRST_YEAR}-{LATEST_YEAR}."
        )
    for year in years:
        m = tables["marriage"][year]["rate_per_1000"]
        d = tables["divorce"][year]["rate_per_1000"]
        if not m > d:
            raise ValueError(f"{year}: marriage rate {m} !> divorce {d}.")
    return {
        "years_covered": [years[0], years[-1]],
        "latest_year": LATEST_YEAR,
        "marriage_gt_divorce_all_years": True,
    }


def build(verbose: bool = True) -> dict[str, Any]:
    fetched_utc = dt.datetime.now(dt.timezone.utc).isoformat()
    if verbose:
        print(f"fetching marriage/divorce: {XLSX_URL}")
    payload = _fetch_bytes(XLSX_URL)
    sha256 = hashlib.sha256(payload).hexdigest()
    tables = parse_workbook(payload)
    validation = validate(tables)
    if verbose:
        for series in ("marriage", "divorce"):
            latest = tables[series][LATEST_YEAR]
            print(
                f"  {series} {LATEST_YEAR}: rate "
                f"{latest['rate_per_1000']}/1,000 "
                f"({latest['count']:,} events)"
            )

    return {
        "schema_version": "nchs_marriage_divorce.v1",
        "latest_year": LATEST_YEAR,
        "first_year": FIRST_YEAR,
        "measure": "crude_rate_per_1000_total_population",
        "report": {
            "title": (
                "National Marriage and Divorce Rate Trends for "
                f"{FIRST_YEAR}-{LATEST_YEAR}"
            ),
            "nvss_citation": NVSS_CITATION,
            "publisher": "National Center for Health Statistics (NCHS)",
            "nvss_page": NVSS_PAGE,
            "faststats_page": FASTATS_PAGE,
            "report_pdf_url": PDF_URL,
        },
        "fetch": {
            "fetched_utc": fetched_utc,
            "fetched_by": "scripts/fetch_nchs_marriage_divorce.py",
            "source_format": "xlsx (cdc.gov/nchs/data/dvs machine-readable)",
            "xlsx_url": XLSX_URL,
            "xlsx_sha256": sha256,
            "n_bytes": len(payload),
        },
        "provisional_note": (
            "NCHS labels this series provisional (permanent designation "
            "for incomplete-reporting-area data), not a signal a final "
            "version exists."
        ),
        "divorce_denominator_note": (
            "The national divorce rate excludes "
            f"{', '.join(DIVORCE_EXCLUDED_STATES)} (they do not report "
            "divorces to NVSS), so its population denominator is smaller "
            "than the marriage rate's. Recorded, not adjusted."
        ),
        "divorce_excluded_states": list(DIVORCE_EXCLUDED_STATES),
        "concept_note": (
            "Crude rates per 1,000 TOTAL population (all ages, both "
            "sexes) -- a loose order-of-magnitude anchor for the PSID "
            "age/duration-specific hazards, never a level target."
        ),
        "tables": {
            series: {str(y): v for y, v in tables[series].items()}
            for series in ("marriage", "divorce")
        },
        "validation": validation,
    }


def main() -> None:
    artifact = build(verbose=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"nchs_marriage_divorce_rates_{LATEST_YEAR}.json"
    out_path.write_text(json.dumps(artifact, indent=2) + "\n")
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
