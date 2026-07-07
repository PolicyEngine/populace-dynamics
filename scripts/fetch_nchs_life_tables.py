"""Fetch and parse the most recent NCHS US period life tables.

The National Center for Health Statistics (NCHS) publishes the
decennial-methodology *United States Life Tables* each year as a
National Vital Statistics Report (NVSR). The most recent complete
report is **United States Life Tables, 2023** (NVSR Vol. 74, No. 6,
released 2025-07-15). NCHS distributes each life table as a small
machine-readable ``.xlsx`` on its FTP server, one file per population:

* ``Table01.xlsx`` -- total population,
* ``Table02.xlsx`` -- male,
* ``Table03.xlsx`` -- female,

(Tables 04-18 are the Hispanic/race-by-sex tables, not fetched here.)
Each sheet is a single-year-of-age period life table with the standard
columns ``qx`` (probability of dying between ages x and x+1), ``lx``
(number surviving to age x, radix 100000), ``dx`` (number dying),
``Lx`` (person-years lived in the age interval), ``Tx`` (total
person-years lived above age x), and ``ex`` (expectation of life at
age x). The final row is the open terminal interval ("100 and older").

This script fetches the three xlsx files, parses each into per-age
rows, validates the standard life-table sanity conditions (radix
``l0 = 100000``; ``ex(0)`` reproduces the report's published headline;
``qx`` non-decreasing across adult ages), and writes one committed
reference JSON with full provenance -- source URLs, the NVSR citation,
the fetch timestamp, the sha256 of each downloaded file, and parse
notes -- to ``data/external/nchs_life_tables_<year>.json``.

The committed JSON is the single source of external mortality truth for
the differential-mortality component (task B1). It is fetched with the
xlsx (not PDF) path because NCHS ships machine-readable spreadsheets;
the PDF report is recorded only as a citation and headline cross-check.

Run from the repository root::

    .venv/bin/python scripts/fetch_nchs_life_tables.py

Re-fetching is idempotent: it overwrites the JSON with the same content
(modulo the fetch timestamp) as long as NCHS has not revised the files.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import json
import re
import ssl
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "external"

# --------------------------------------------------------------------------
# Vintage and source registry (the one place a new NCHS release is pinned).
# --------------------------------------------------------------------------
#: Most recent complete NCHS United States Life Tables vintage.
VINTAGE_YEAR = 2023

#: NVSR volume/number for the vintage; both the report PDF and the FTP
#: xlsx directory key off ``<vol>-<num>``.
NVSR_VOLUME = 74
NVSR_NUMBER = 6

_FTP_BASE = (
    "https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/NVSR/"
    f"{NVSR_VOLUME}-{NVSR_NUMBER:02d}"
)
REPORT_PDF_URL = (
    f"https://www.cdc.gov/nchs/data/nvsr/nvsr{NVSR_VOLUME}/"
    f"nvsr{NVSR_VOLUME}-{NVSR_NUMBER:02d}.pdf"
)
PRODUCTS_PAGE = "https://www.cdc.gov/nchs/products/life_tables.htm"

#: population key -> (FTP xlsx filename, expected sheet-title fragment).
#: The title fragment is verified against the parsed sheet so a shifted
#: NCHS layout (e.g. a Table renumbering) fails loudly rather than
#: silently loading the wrong population.
SOURCES: dict[str, tuple[str, str]] = {
    "total": ("Table01.xlsx", "total population"),
    "male": ("Table02.xlsx", "males"),
    "female": ("Table03.xlsx", "females"),
}

#: The six numeric life-table columns, in the sheet's column order after
#: the age label. The parser verifies the short header row matches this.
LIFE_TABLE_COLUMNS = ("qx", "lx", "dx", "Lx", "Tx", "ex")

#: Published headline life expectancy at birth, transcribed from the
#: NVSR 74-6 report (2023) for the cross-check. ex(0) parsed from each
#: xlsx must round to these to one decimal, or the parse is rejected.
PUBLISHED_EX0 = {"total": 78.4, "male": 75.8, "female": 81.1}

#: Adult age from which qx must be non-decreasing (below it the
#: young-adult "accident hump" makes qx non-monotone by design).
QX_MONOTONE_FROM_AGE = 40


# --------------------------------------------------------------------------
# Fetch
# --------------------------------------------------------------------------
def _ssl_context() -> ssl.SSLContext:
    """A verifying SSL context, preferring certifi's CA bundle.

    Standalone CPython builds (e.g. the uv-managed interpreter) do not
    always trust the system CA store, so a plain default context can
    fail to verify ``ftp.cdc.gov``. Using certifi's bundle when present
    keeps verification ON while making the fetch portable.
    """
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl.create_default_context()


def _fetch_bytes(url: str) -> bytes:
    """Download ``url`` and return the raw bytes (browser-like UA)."""
    request = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (populace-dynamics fetch)"}
    )
    with urllib.request.urlopen(
        request, timeout=120, context=_ssl_context()
    ) as response:
        return response.read()


def _parse_age(label: str) -> int:
    """Leading integer age from a life-table row label.

    Handles ``"0-1"``, ``"0–1"`` (en-dash), and the open terminal
    ``"100 and older"``; returns the first year of the interval.
    """
    match = re.match(r"\s*(\d+)", label)
    if not match:
        raise ValueError(f"Cannot parse an age from row label {label!r}")
    return int(match.group(1))


def parse_life_table(
    xlsx_bytes: bytes, expected_title_fragment: str
) -> tuple[list[dict[str, Any]], str]:
    """Parse one NCHS life-table xlsx into per-age rows.

    Returns ``(rows, table_title)`` where each row is
    ``{"age", "qx", "lx", "dx", "Lx", "Tx", "ex"}`` in ascending age,
    and ``table_title`` is the sheet's title cell (A1). The short header
    row (``qx``/``lx``/.../``ex``) and the title-cell population fragment
    are both verified, so a changed NCHS layout raises rather than
    silently mislabeling.
    """
    workbook = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    sheet = workbook[workbook.sheetnames[0]]
    rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    workbook.close()

    title = str(rows[0][0] or "").strip()
    if expected_title_fragment.lower() not in title.lower():
        raise ValueError(
            f"Sheet title {title!r} does not contain the expected "
            f"population fragment {expected_title_fragment!r}; the NCHS "
            "layout may have changed."
        )
    if str(VINTAGE_YEAR) not in title:
        raise ValueError(
            f"Sheet title {title!r} does not carry the expected vintage "
            f"year {VINTAGE_YEAR}."
        )

    # Row index 2 (third row) is the short header: (None, qx, lx, ...).
    short_header = tuple(
        str(c).strip() if c is not None else None for c in rows[2]
    )
    if short_header[1 : 1 + len(LIFE_TABLE_COLUMNS)] != LIFE_TABLE_COLUMNS:
        raise ValueError(
            f"Unexpected life-table header {short_header!r}; expected "
            f"columns {LIFE_TABLE_COLUMNS} after the age label."
        )

    parsed: list[dict[str, Any]] = []
    for raw in rows[3:]:
        if raw[0] is None:
            continue
        label = str(raw[0]).strip()
        if not re.match(r"\s*\d", label):  # SOURCE / footnote lines
            continue
        age = _parse_age(label)
        record: dict[str, Any] = {"age": age, "age_label": label}
        for offset, name in enumerate(LIFE_TABLE_COLUMNS, start=1):
            value = raw[offset]
            record[name] = None if value is None else float(value)
        parsed.append(record)

    parsed.sort(key=lambda r: r["age"])
    return parsed, title


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------
def validate_table(
    population: str, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    """Standard life-table sanity checks; returns the recorded results.

    Raises ``ValueError`` on any hard failure (radix, coverage, ex(0)
    headline, adult qx monotonicity), so a corrupt fetch never reaches
    the committed artifact.
    """
    ages = [r["age"] for r in rows]
    if ages != list(range(ages[0], ages[-1] + 1)):
        raise ValueError(
            f"{population}: ages are not contiguous single years "
            f"({ages[0]}..{ages[-1]})."
        )
    if ages[0] != 0:
        raise ValueError(f"{population}: life table does not start at age 0.")

    l0 = rows[0]["lx"]
    if round(l0) != 100000:
        raise ValueError(f"{population}: radix l0={l0} is not 100000.")

    ex0 = rows[0]["ex"]
    published = PUBLISHED_EX0[population]
    if round(ex0, 1) != published:
        raise ValueError(
            f"{population}: parsed ex(0)={ex0:.4f} does not match the "
            f"published headline {published}."
        )

    # qx non-decreasing across adult ages (below QX_MONOTONE_FROM_AGE the
    # accident hump makes it non-monotone by design); terminal qx == 1.
    by_age = {r["age"]: r for r in rows}
    violations = []
    for age in range(QX_MONOTONE_FROM_AGE, ages[-1]):
        if by_age[age]["qx"] > by_age[age + 1]["qx"] + 1e-9:
            violations.append(age)
    if violations:
        raise ValueError(
            f"{population}: qx decreases at adult age(s) {violations}."
        )
    terminal_qx = rows[-1]["qx"]
    if round(terminal_qx, 6) != 1.0:
        raise ValueError(
            f"{population}: terminal qx={terminal_qx} is not 1 "
            f"(open interval expected at age {ages[-1]})."
        )

    return {
        "l0": l0,
        "ex0": ex0,
        "ex0_published_headline": published,
        "terminal_age": ages[-1],
        "n_ages": len(rows),
        "qx_nondecreasing_from_age": QX_MONOTONE_FROM_AGE,
    }


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------
def build(verbose: bool = True) -> dict[str, Any]:
    """Fetch, parse, validate the three tables; return the artifact dict."""
    fetched_utc = dt.datetime.now(dt.timezone.utc).isoformat()
    tables: dict[str, list[dict[str, Any]]] = {}
    source_files: dict[str, dict[str, Any]] = {}
    validation: dict[str, Any] = {}
    headline_ex0: dict[str, float] = {}

    for population, (filename, title_fragment) in SOURCES.items():
        url = f"{_FTP_BASE}/{filename}"
        if verbose:
            print(f"fetching {population}: {url}")
        payload = _fetch_bytes(url)
        sha256 = hashlib.sha256(payload).hexdigest()
        rows, table_title = parse_life_table(payload, title_fragment)
        validation[population] = validate_table(population, rows)
        headline_ex0[population] = round(rows[0]["ex"], 4)
        tables[population] = rows
        source_files[population] = {
            "url": url,
            "filename": filename,
            "sheet_title": table_title,
            "sha256": sha256,
            "n_bytes": len(payload),
        }
        if verbose:
            print(
                f"  {population}: {len(rows)} ages, ex(0)="
                f"{rows[0]['ex']:.2f}, sha256={sha256[:12]}..."
            )

    return {
        "schema_version": "nchs_life_tables.v1",
        "vintage_year": VINTAGE_YEAR,
        "report": {
            "title": f"United States Life Tables, {VINTAGE_YEAR}",
            "nvsr_citation": (
                "Arias E, Xu J, Kochanek K. United States Life Tables, "
                f"{VINTAGE_YEAR}. National Vital Statistics Reports, "
                f"Vol. {NVSR_VOLUME}, No. {NVSR_NUMBER}. Hyattsville, MD: "
                "National Center for Health Statistics; July 15, 2025."
            ),
            "publisher": "National Center for Health Statistics (NCHS)",
            "report_pdf_url": REPORT_PDF_URL,
            "products_page": PRODUCTS_PAGE,
        },
        "fetch": {
            "fetched_utc": fetched_utc,
            "fetched_by": "scripts/fetch_nchs_life_tables.py",
            "source_format": "xlsx (NCHS FTP machine-readable spreadsheet)",
            "source_files": source_files,
        },
        "parse_notes": (
            "Parsed with openpyxl from the NCHS FTP xlsx (one file per "
            "population): row 1 title, row 2 long headers, row 3 short "
            "headers (qx, lx, dx, Lx, Tx, ex), rows 4+ single-year-of-age "
            "data ending at the open terminal interval '100 and older' "
            "(age 100). Age is the first year of each interval. The sheet "
            "title's population fragment and vintage year, and the short "
            "header row, are verified at parse; l0=100000, ex(0) vs the "
            "published headline, and adult qx monotonicity are validated. "
            "Numeric values are the xlsx-stored doubles (NCHS stores lx/dx/"
            "Lx/Tx at float precision, so lx(0) reads as exactly 100000 "
            "but interior lx carry fractional parts)."
        ),
        "columns": ["age", "age_label", *LIFE_TABLE_COLUMNS],
        "terminal_age": tables["total"][-1]["age"],
        "headline_ex0": headline_ex0,
        "validation": validation,
        "tables": tables,
    }


def main() -> None:
    artifact = build(verbose=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"nchs_life_tables_{VINTAGE_YEAR}.json"
    out_path.write_text(json.dumps(artifact, indent=2) + "\n")
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
