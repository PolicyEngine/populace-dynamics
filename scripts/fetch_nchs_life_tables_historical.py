"""Fetch two historical NCHS US period life-table vintages (2000, 2010).

The gate-2 candidate-5 registration (issue #42 comment 4911788302) fixes
the spouse-death mortality *trend* from EXTERNAL NCHS life tables rather
than the train-fit Poisson slope of candidate 4: the per-sex log-linear
period slope ``beta_sex`` is the slope of the age-45-84-band-average
``qx`` across three NCHS United States Life Table vintages. Candidate 4's
committed reference already carries the most recent vintage
(``data/external/nchs_life_tables_2023.json``, NVSR 74-6); this script
fetches the two additional historical vintages the registration targets --
the decennial-era **United States Life Tables, 2000** and the mid-panel
**United States Life Tables, 2010** -- and commits them in the SAME
``nchs_life_tables.v1`` schema with full provenance.

Two source formats, each the most machine-readable NCHS ships for its
vintage:

* **2010** -- NVSR Vol. 63, No. 7 (Arias E., released 2014-11-06). NCHS
  distributes the per-population single-year-of-age life table as a small
  ``.xlsx`` on its FTP server (the "Spreadsheet version" the report's own
  table header points to), under ``NVSR/63_07/Table0{1,2,3}.xlsx`` -- the
  identical layout to the 2023 xlsx, parsed the same way.
* **2000** -- NVSR Vol. 51, No. 3 (Arias E., released 2002-12-19). This
  2002 report predates the machine-readable spreadsheet distribution
  (no xlsx on the FTP server), so the single-year-of-age life table is
  parsed from the published NVSR PDF via ``pdftotext -layout`` (poppler):
  Table 1 (total), Table 2 (males), Table 3 (females). The PDF is the
  authoritative published report -- the same document the 2023 artifact
  records as ``report_pdf_url`` provenance.

Each parsed table is validated with the standard life-table sanity
conditions (radix ``l0 = 100000``; ``ex(0)`` reproduces the report's
published headline, transcribed from the report ABSTRACT independently of
the detailed table; adult ``qx`` non-decreasing from age 40; open terminal
interval ``qx = 1``) before it can reach the committed JSON, so a mis-parse
fails loudly. Written to ``data/external/nchs_life_tables_<year>.json``.

Run from the repository root (needs ``openpyxl`` for the 2010 xlsx and
``pdftotext`` on PATH for the 2000 PDF)::

    .venv/bin/python scripts/fetch_nchs_life_tables_historical.py

Re-fetching is idempotent modulo the fetch timestamp as long as NCHS has
not revised the files.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import json
import re
import ssl
import subprocess
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "external"

#: The six numeric life-table columns, in the sheet/table column order
#: after the age label (identical across vintages and source formats).
LIFE_TABLE_COLUMNS = ("qx", "lx", "dx", "Lx", "Tx", "ex")

#: Adult age from which qx must be non-decreasing (below it the young-adult
#: accident hump makes qx non-monotone by design). Matches the 2023 fetch.
QX_MONOTONE_FROM_AGE = 40

#: Per-population xlsx table filenames (2010 FTP; identical to 2023).
XLSX_TABLES = {
    "total": ("Table01.xlsx", "total population"),
    "male": ("Table02.xlsx", "males"),
    "female": ("Table03.xlsx", "females"),
}
#: Per-population PDF table numbers (2000 NVSR report).
PDF_TABLES = {
    "total": (1, "total population"),
    "male": (2, "males"),
    "female": (3, "females"),
}

# --------------------------------------------------------------------------
# Vintage registry (the one place a historical NCHS vintage is pinned).
# ``ex0_published`` is transcribed from each report's ABSTRACT / summary
# table (independently of the detailed life table), so the parse's ex(0) is
# cross-checked against a source OTHER than the table it parses.
# --------------------------------------------------------------------------
VINTAGES: dict[int, dict[str, Any]] = {
    2000: {
        "format": "pdf",
        "nvsr_volume": 51,
        "nvsr_number": 3,
        "released": "December 19, 2002",
        "report_pdf_url": (
            "https://www.cdc.gov/nchs/data/nvsr/nvsr51/nvsr51_03.pdf"
        ),
        "nvsr_citation": (
            "Arias E. United States Life Tables, 2000. National Vital "
            "Statistics Reports, Vol. 51, No. 3. Hyattsville, MD: National "
            "Center for Health Statistics; December 19, 2002."
        ),
        "ex0_published": {"total": 76.9, "male": 74.1, "female": 79.5},
    },
    2010: {
        "format": "xlsx",
        "nvsr_volume": 63,
        "nvsr_number": 7,
        "released": "November 6, 2014",
        "ftp_dir": (
            "https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/"
            "NVSR/63_07"
        ),
        "report_pdf_url": (
            "https://www.cdc.gov/nchs/data/nvsr/nvsr63/nvsr63_07.pdf"
        ),
        "nvsr_citation": (
            "Arias E. United States Life Tables, 2010. National Vital "
            "Statistics Reports, Vol. 63, No. 7. Hyattsville, MD: National "
            "Center for Health Statistics; November 6, 2014."
        ),
        "ex0_published": {"total": 78.7, "male": 76.2, "female": 81.0},
    },
}

PRODUCTS_PAGE = "https://www.cdc.gov/nchs/products/life_tables.htm"


# --------------------------------------------------------------------------
# Fetch
# --------------------------------------------------------------------------
def _ssl_context() -> ssl.SSLContext:
    """A verifying SSL context, preferring certifi's CA bundle."""
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl.create_default_context()


def _fetch_bytes(url: str) -> bytes:
    request = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (populace-dynamics fetch)"}
    )
    with urllib.request.urlopen(
        request, timeout=120, context=_ssl_context()
    ) as response:
        return response.read()


def _parse_age(label: str) -> int:
    """Leading integer age from a life-table row label.

    Handles ``"0-1"``, ``"0-1"`` (en-dash), ``"100 and over"`` and
    ``"100 years and over"``; returns the first year of the interval.
    """
    match = re.match(r"\s*(\d+)", label)
    if not match:
        raise ValueError(f"Cannot parse an age from row label {label!r}")
    return int(match.group(1))


# --------------------------------------------------------------------------
# xlsx parse (2010; identical layout to the committed 2023 reference)
# --------------------------------------------------------------------------
def _norm_header(cell: Any) -> str:
    """Normalise a header cell: strip and drop parentheses (``q(x)`` -> ``qx``).

    NCHS ships the short header as ``qx``/``lx`` in the 2023 xlsx but as
    ``q(x)``/``l(x)`` in the 2010 xlsx; both normalise to the same tokens.
    """
    if cell is None:
        return ""
    return re.sub(r"[()]", "", str(cell).strip())


def parse_life_table_xlsx(
    xlsx_bytes: bytes, expected_title_fragment: str, vintage_year: int
) -> tuple[list[dict[str, Any]], str]:
    """Parse one NCHS life-table xlsx into per-age rows (vintage-parametrised).

    Row 1 is the sheet title; the short header row (whose second column
    normalises to ``qx`` and third to ``lx``) is located dynamically -- the
    2023 xlsx carries it at row 3 as ``qx``/``lx``, the 2010 xlsx at row 7
    as ``q(x)``/``l(x)`` after a taller multi-line header -- and the
    single-year-of-age data follow it. The population fragment and vintage
    year in the title, and the six normalised column labels, are verified so
    a changed NCHS layout raises rather than silently mislabeling.
    """
    workbook = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), data_only=True, read_only=True
    )
    sheet = workbook[workbook.sheetnames[0]]
    rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    workbook.close()

    title = str(rows[0][0] or "").strip()
    if expected_title_fragment.lower() not in title.lower():
        raise ValueError(
            f"Sheet title {title!r} does not contain the expected "
            f"population fragment {expected_title_fragment!r}."
        )
    if str(vintage_year) not in title:
        raise ValueError(
            f"Sheet title {title!r} does not carry the expected vintage "
            f"year {vintage_year}."
        )

    header_idx = None
    for i, raw in enumerate(rows):
        if (
            len(raw) > len(LIFE_TABLE_COLUMNS)
            and _norm_header(raw[1]).lower() == "qx"
            and _norm_header(raw[2]).lower() == "lx"
        ):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "No life-table header row (a row with qx/lx columns) found."
        )
    header = tuple(
        _norm_header(c)
        for c in rows[header_idx][1 : 1 + len(LIFE_TABLE_COLUMNS)]
    )
    if header != LIFE_TABLE_COLUMNS:
        raise ValueError(
            f"Unexpected life-table header {header!r}; expected columns "
            f"{LIFE_TABLE_COLUMNS} after the age label."
        )

    parsed: list[dict[str, Any]] = []
    for raw in rows[header_idx + 1 :]:
        if raw[0] is None:
            continue
        label = str(raw[0]).strip()
        if not re.match(r"\s*\d", label):
            continue
        age = _parse_age(label)
        record: dict[str, Any] = {"age": age, "age_label": label}
        for offset, name in enumerate(LIFE_TABLE_COLUMNS, start=1):
            value = raw[offset]
            record[name] = None if value is None else float(value)
        parsed.append(record)

    parsed.sort(key=lambda r: r["age"])
    return parsed, title


# --------------------------------------------------------------------------
# PDF parse (2000; NVSR report via pdftotext -layout)
# --------------------------------------------------------------------------
def _pdftotext_layout(pdf_bytes: bytes) -> str:
    """Render a PDF to layout-preserving text via poppler ``pdftotext``."""
    proc = subprocess.run(
        ["pdftotext", "-layout", "-", "-"],
        input=pdf_bytes,
        capture_output=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"pdftotext failed (rc={proc.returncode}): "
            f"{proc.stderr.decode(errors='replace')[:400]}"
        )
    return proc.stdout.decode("utf-8", errors="replace")


_NUM_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")


def parse_life_table_pdf(
    text: str, table_number: int, vintage_year: int, expected_title: str
) -> tuple[list[dict[str, Any]], str]:
    """Parse one single-year life table from NVSR ``pdftotext -layout`` text.

    The section for population table ``table_number`` runs from its first
    title line ``"Table N. Life table for <pop>: United States, <year>"``
    (the "-Con." continuation shares the title) up to the first title line
    of table ``N + 1``. Within the section, a data row is any line whose
    leading integer equals the next expected age and whose last six numeric
    tokens are (qx, lx, dx, Lx, Tx, ex) with ``0 < qx <= 1`` and
    ``lx <= 100000`` -- the age-sequence + qx/lx guards reject repeated
    column headers, page headers, and footers. Returns ``(rows, title)``.
    """
    lines = text.splitlines()
    title_re = re.compile(
        rf"^\s*Table\s+(\d+)\.\s*Life table for .*:\s*United States,\s*"
        rf"{vintage_year}"
    )
    first_idx: dict[int, int] = {}
    for i, line in enumerate(lines):
        m = title_re.match(line)
        if m:
            n = int(m.group(1))
            first_idx.setdefault(n, i)

    if table_number not in first_idx:
        raise ValueError(
            f"Table {table_number} title not found for vintage "
            f"{vintage_year}."
        )
    start = first_idx[table_number]
    later = [i for n, i in first_idx.items() if n > table_number]
    end = min(later) if later else len(lines)

    section_title = lines[start].strip()
    if expected_title.lower() not in section_title.lower():
        raise ValueError(
            f"Table {table_number} title {section_title!r} does not contain "
            f"{expected_title!r}."
        )

    parsed: list[dict[str, Any]] = []
    expected_age = 0
    for line in lines[start:end]:
        lead = re.match(r"\s*(\d+)", line)
        if not lead:
            continue
        age = int(lead.group(1))
        if age != expected_age:
            continue
        nums = _NUM_RE.findall(line)
        if len(nums) < 7:  # >= 1 age int + 6 data columns
            continue
        cols = [float(tok.replace(",", "")) for tok in nums[-6:]]
        qx, lx = cols[0], cols[1]
        if not (0.0 < qx <= 1.0000001) or lx > 100000.5:
            continue
        label = line.strip().split(" .", 1)[0].strip()
        record: dict[str, Any] = {"age": age, "age_label": label}
        for name, value in zip(LIFE_TABLE_COLUMNS, cols, strict=True):
            record[name] = value
        parsed.append(record)
        expected_age += 1

    parsed.sort(key=lambda r: r["age"])
    return parsed, section_title


# --------------------------------------------------------------------------
# Validation (shared)
# --------------------------------------------------------------------------
def validate_table(
    population: str,
    rows: list[dict[str, Any]],
    ex0_published: float,
) -> dict[str, Any]:
    """Standard life-table sanity checks; returns the recorded results."""
    ages = [r["age"] for r in rows]
    if ages != list(range(ages[0], ages[-1] + 1)):
        raise ValueError(
            f"{population}: ages are not contiguous single years "
            f"({ages[0]}..{ages[-1]})."
        )
    if ages[0] != 0:
        raise ValueError(f"{population}: life table does not start at age 0.")

    l0 = rows[0]["lx"]
    if round(l0) != 100000:
        raise ValueError(f"{population}: radix l0={l0} is not 100000.")

    ex0 = rows[0]["ex"]
    if round(ex0, 1) != ex0_published:
        raise ValueError(
            f"{population}: parsed ex(0)={ex0:.4f} does not match the "
            f"published headline {ex0_published}."
        )

    by_age = {r["age"]: r for r in rows}
    violations = [
        age
        for age in range(QX_MONOTONE_FROM_AGE, ages[-1])
        if by_age[age]["qx"] > by_age[age + 1]["qx"] + 1e-9
    ]
    if violations:
        raise ValueError(
            f"{population}: qx decreases at adult age(s) {violations}."
        )
    terminal_qx = rows[-1]["qx"]
    if round(terminal_qx, 5) != 1.0:
        raise ValueError(
            f"{population}: terminal qx={terminal_qx} is not 1 "
            f"(open interval expected at age {ages[-1]})."
        )

    return {
        "l0": l0,
        "ex0": ex0,
        "ex0_published_headline": ex0_published,
        "terminal_age": ages[-1],
        "n_ages": len(rows),
        "qx_nondecreasing_from_age": QX_MONOTONE_FROM_AGE,
    }


# --------------------------------------------------------------------------
# Drivers per source format
# --------------------------------------------------------------------------
def build_xlsx_vintage(
    vintage_year: int, cfg: dict[str, Any], verbose: bool
) -> dict[str, Any]:
    fetched_utc = dt.datetime.now(dt.timezone.utc).isoformat()
    tables: dict[str, list[dict[str, Any]]] = {}
    source_files: dict[str, dict[str, Any]] = {}
    validation: dict[str, Any] = {}
    headline_ex0: dict[str, float] = {}
    for population, (filename, fragment) in XLSX_TABLES.items():
        url = f"{cfg['ftp_dir']}/{filename}"
        if verbose:
            print(f"fetching {vintage_year} {population}: {url}")
        payload = _fetch_bytes(url)
        sha256 = hashlib.sha256(payload).hexdigest()
        rows, title = parse_life_table_xlsx(payload, fragment, vintage_year)
        validation[population] = validate_table(
            population, rows, cfg["ex0_published"][population]
        )
        headline_ex0[population] = round(rows[0]["ex"], 4)
        tables[population] = rows
        source_files[population] = {
            "url": url,
            "filename": filename,
            "sheet_title": title,
            "sha256": sha256,
            "n_bytes": len(payload),
        }
    parse_notes = (
        "Parsed with openpyxl from the NCHS FTP xlsx (one file per "
        "population; identical layout to the committed 2023 reference): "
        "row 1 title, row 3 short headers (qx, lx, dx, Lx, Tx, ex), rows 4+ "
        "single-year-of-age data to the open terminal interval (age 100). "
        "The title's population fragment and vintage year and the short "
        "header are verified at parse; l0=100000, ex(0) vs the published "
        "headline, and adult qx monotonicity are validated."
    )
    return _assemble(
        vintage_year,
        cfg,
        fetched_utc,
        "xlsx (NCHS FTP machine-readable spreadsheet)",
        source_files,
        parse_notes,
        tables,
        headline_ex0,
        validation,
    )


def build_pdf_vintage(
    vintage_year: int, cfg: dict[str, Any], verbose: bool
) -> dict[str, Any]:
    fetched_utc = dt.datetime.now(dt.timezone.utc).isoformat()
    url = cfg["report_pdf_url"]
    filename = url.rsplit("/", 1)[-1]
    if verbose:
        print(f"fetching {vintage_year} report PDF: {url}")
    payload = _fetch_bytes(url)
    sha256 = hashlib.sha256(payload).hexdigest()
    text = _pdftotext_layout(payload)

    tables: dict[str, list[dict[str, Any]]] = {}
    source_files: dict[str, dict[str, Any]] = {}
    validation: dict[str, Any] = {}
    headline_ex0: dict[str, float] = {}
    for population, (table_number, fragment) in PDF_TABLES.items():
        expected_title = (
            f"Life table for the {fragment}"
            if (population == "total")
            else f"Life table for {fragment}"
        )
        rows, title = parse_life_table_pdf(
            text, table_number, vintage_year, expected_title
        )
        validation[population] = validate_table(
            population, rows, cfg["ex0_published"][population]
        )
        headline_ex0[population] = round(rows[0]["ex"], 4)
        tables[population] = rows
        source_files[population] = {
            "url": url,
            "filename": filename,
            "pdf_table_number": table_number,
            "pdf_table_title": title,
            "sha256": sha256,
            "n_bytes": len(payload),
        }
        if verbose:
            print(
                f"  {population}: {len(rows)} ages, ex(0)={rows[0]['ex']:.2f}"
            )
    parse_notes = (
        "Parsed from the published NVSR report PDF via 'pdftotext -layout' "
        "(poppler): Table 1 (total population), Table 2 (males), Table 3 "
        "(females), each a single-year-of-age life table to the open "
        "terminal interval '100 years and over' (age 100). This 2002 report "
        "predates NCHS's machine-readable spreadsheet distribution (no xlsx "
        "on the FTP server), so the authoritative published PDF is the "
        "source. A data row is the line whose leading integer is the next "
        "expected age and whose last six numeric tokens are (qx, lx, dx, "
        "Lx, Tx, ex) with 0 < qx <= 1 and lx <= 100000 -- the age-sequence "
        "and qx/lx guards reject repeated headers and page footers. "
        "l0=100000, ex(0) vs the abstract-transcribed published headline, "
        "and adult qx monotonicity are validated per population."
    )
    return _assemble(
        vintage_year,
        cfg,
        fetched_utc,
        "pdf (NVSR report, pdftotext -layout)",
        source_files,
        parse_notes,
        tables,
        headline_ex0,
        validation,
    )


def _assemble(
    vintage_year: int,
    cfg: dict[str, Any],
    fetched_utc: str,
    source_format: str,
    source_files: dict[str, Any],
    parse_notes: str,
    tables: dict[str, list[dict[str, Any]]],
    headline_ex0: dict[str, float],
    validation: dict[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": "nchs_life_tables.v1",
        "vintage_year": vintage_year,
        "report": {
            "title": f"United States Life Tables, {vintage_year}",
            "nvsr_citation": cfg["nvsr_citation"],
            "publisher": "National Center for Health Statistics (NCHS)",
            "nvsr_volume": cfg["nvsr_volume"],
            "nvsr_number": cfg["nvsr_number"],
            "released": cfg["released"],
            "report_pdf_url": cfg["report_pdf_url"],
            "products_page": PRODUCTS_PAGE,
        },
        "fetch": {
            "fetched_utc": fetched_utc,
            "fetched_by": "scripts/fetch_nchs_life_tables_historical.py",
            "source_format": source_format,
            "source_files": source_files,
        },
        "parse_notes": parse_notes,
        "columns": ["age", "age_label", *LIFE_TABLE_COLUMNS],
        "terminal_age": tables["total"][-1]["age"],
        "headline_ex0": headline_ex0,
        "validation": validation,
        "tables": tables,
    }


def build(vintage_year: int, verbose: bool = True) -> dict[str, Any]:
    cfg = VINTAGES[vintage_year]
    if cfg["format"] == "xlsx":
        return build_xlsx_vintage(vintage_year, cfg, verbose)
    return build_pdf_vintage(vintage_year, cfg, verbose)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for vintage_year in sorted(VINTAGES):
        artifact = build(vintage_year, verbose=True)
        out_path = OUT_DIR / f"nchs_life_tables_{vintage_year}.json"
        out_path.write_text(json.dumps(artifact, indent=2) + "\n")
        print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
